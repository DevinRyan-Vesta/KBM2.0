# inventory/import_views.py
"""Import functionality for inventory items"""
from flask import render_template, request, redirect, url_for, flash, jsonify, session
from flask_login import login_required, current_user
import csv
import difflib
import io
from collections import OrderedDict
from typing import List, Dict, Any, Optional, Tuple
from werkzeug.utils import secure_filename

from utilities.tenant_helpers import tenant_query, tenant_add, tenant_commit, tenant_rollback, get_tenant_session
from middleware.tenant_middleware import tenant_required
from utilities.database import db, Item, Property, PropertyUnit, utc_now
from . import inventory_bp


ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

# Max rows we'll keep in the session for an import. Beyond this we slice +
# warn the user; sessions get expensive past a few hundred KB.
MAX_IMPORT_ROWS = 1000

# difflib similarity cutoff for fuzzy property/unit name matches. 0.85 keeps
# typos ("Smith Apartment" vs "Smith Apartments") and rejects abbreviations
# ("Smith Apts"). Tighter is safer for incorrect assignments.
FUZZY_MATCH_CUTOFF = 0.85

# Cap on per-row warnings flashed at the end of an import. Anything beyond
# this is summarized as "...and N more".
MAX_WARNINGS_DISPLAYED = 10


# ---- Fuzzy matching helpers --------------------------------------------------

def _normalize(s) -> str:
    """Trim + lowercase. Used for case-insensitive equality before fuzzy matching."""
    if s is None:
        return ''
    return str(s).strip().lower()


def _fuzzy_find_property(name: str, all_properties) -> Tuple[Optional[Property], bool]:
    """Find a Property by name. Returns (property, is_fuzzy_match).

    Tries exact case-insensitive match (after stripping) first, falls back to
    fuzzy match against the cutoff. Returns (None, False) if no match.
    """
    target = _normalize(name)
    if not target:
        return None, False

    by_norm_name = {}
    for prop in all_properties:
        key = _normalize(prop.name)
        if key and key not in by_norm_name:
            by_norm_name[key] = prop

    exact = by_norm_name.get(target)
    if exact is not None:
        return exact, False

    matches = difflib.get_close_matches(target, list(by_norm_name.keys()), n=1, cutoff=FUZZY_MATCH_CUTOFF)
    if matches:
        return by_norm_name[matches[0]], True
    return None, False


def _fuzzy_find_unit(label: str, units) -> Tuple[Optional[PropertyUnit], bool]:
    """Find a PropertyUnit by label within a given property's units."""
    target = _normalize(label)
    if not target:
        return None, False

    by_norm_label = {}
    for unit in units:
        key = _normalize(unit.label)
        if key and key not in by_norm_label:
            by_norm_label[key] = unit

    exact = by_norm_label.get(target)
    if exact is not None:
        return exact, False

    matches = difflib.get_close_matches(target, list(by_norm_label.keys()), n=1, cutoff=FUZZY_MATCH_CUTOFF)
    if matches:
        return by_norm_label[matches[0]], True
    return None, False


def _similarity_pct(a: str, b: str) -> int:
    """Rough integer percent similarity for UI display."""
    return int(difflib.SequenceMatcher(None, _normalize(a), _normalize(b)).ratio() * 100)


def _flash_warning_summary(warnings: List[str]) -> None:
    """Flash up to MAX_WARNINGS_DISPLAYED individual warnings + a summary."""
    if not warnings:
        return
    visible = warnings[:MAX_WARNINGS_DISPLAYED]
    extra = len(warnings) - len(visible)
    for w in visible:
        flash(w, "warning")
    if extra > 0:
        flash(f"...and {extra} more import warning(s) suppressed.", "warning")


# ---- Import analysis (pre-resolution) ----------------------------------------

def _analyze_import(rows: List[Dict[str, str]], mapping: Dict[str, str], item_type_canonical: str) -> Dict[str, Any]:
    """Inspect a parsed file to power the resolve/review wizard.

    Performs:
      - In-file dedup by label (first occurrence wins; subsequent dropped).
      - Cross-tenant duplicate check by label against existing items of this type.
      - Property name classification (exact / fuzzy / missing).

    Returns a dict of summarized data the resolve template renders. Unit
    handling is deferred to the process step (where it's gated on the
    auto-create-units checkbox + whether the resolved property is brand new).

    item_type_canonical: 'Key' | 'Lockbox' | 'Sign' (matches Item.type values)
    """
    label_col = mapping.get('label')
    property_col = mapping.get('property_name')
    address_col = mapping.get('address')

    # In-file dedup by label
    seen_labels = {}
    deduped = []
    in_file_dups = []  # list of (row_idx, label, dropped_because_dup_of_row)

    for idx, row in enumerate(rows, 1):
        label = (row.get(label_col, '') or '').strip() if label_col else ''
        if not label:
            # blank labels handled in process step (counted as errors there)
            deduped.append((idx, row))
            continue
        norm = _normalize(label)
        first_idx = seen_labels.get(norm)
        if first_idx is not None:
            in_file_dups.append((idx, label, first_idx))
            continue
        seen_labels[norm] = idx
        deduped.append((idx, row))

    # Existing tenant items of this type — for duplicate detection
    existing_items = tenant_query(Item).filter(Item.type == item_type_canonical).all()
    existing_by_label = {}
    for item in existing_items:
        key = _normalize(item.label)
        if key and key not in existing_by_label:
            existing_by_label[key] = item

    duplicates = []
    for idx, row in deduped:
        label = (row.get(label_col, '') or '').strip() if label_col else ''
        if not label:
            continue
        existing = existing_by_label.get(_normalize(label))
        if existing:
            duplicates.append({
                'row_idx': idx,
                'label': label,
                'existing_id': existing.id,
                'existing_label': existing.label,
            })

    # Properties classification
    all_properties = tenant_query(Property).all()
    property_options = sorted([(p.id, p.name) for p in all_properties], key=lambda x: x[1].lower())

    unique_property_names = OrderedDict()
    if property_col:
        for idx, row in deduped:
            name = (row.get(property_col, '') or '').strip()
            if not name:
                continue
            if name not in unique_property_names:
                addr = (row.get(address_col, '') or '').strip() if address_col else ''
                unique_property_names[name] = {
                    'sample_address': addr or None,
                    'first_row_idx': idx,
                }

    property_resolution = []
    for name, meta in unique_property_names.items():
        prop, is_fuzzy = _fuzzy_find_property(name, all_properties)
        entry = {
            'name': name,
            'sample_address': meta['sample_address'],
            'first_row_idx': meta['first_row_idx'],
        }
        if prop and not is_fuzzy:
            entry['status'] = 'exact'
            entry['matched_property_id'] = prop.id
            entry['matched_property_name'] = prop.name
        elif prop and is_fuzzy:
            entry['status'] = 'fuzzy'
            entry['matched_property_id'] = prop.id
            entry['matched_property_name'] = prop.name
            entry['similarity'] = _similarity_pct(name, prop.name)
        else:
            entry['status'] = 'missing'
            entry['matched_property_id'] = None
            entry['matched_property_name'] = None
        property_resolution.append(entry)

    needs_property_review = sum(1 for p in property_resolution if p['status'] in ('fuzzy', 'missing'))
    exact_property_matches = sum(1 for p in property_resolution if p['status'] == 'exact')

    # "Ready" rows: not duplicate, and either no property_name or exact match
    duplicate_row_idxs = {d['row_idx'] for d in duplicates}
    needs_review_property_names = {
        p['name'] for p in property_resolution if p['status'] in ('fuzzy', 'missing')
    }
    ready_count = 0
    for idx, row in deduped:
        if idx in duplicate_row_idxs:
            continue
        prop_name = (row.get(property_col, '') or '').strip() if property_col else ''
        if prop_name and prop_name in needs_review_property_names:
            continue
        ready_count += 1

    return {
        'rows': deduped,
        'in_file_dups': in_file_dups,
        'duplicates': duplicates,
        'properties': property_resolution,
        'property_options': property_options,
        'stats': {
            'total_rows_in_file': len(rows),
            'in_file_dup_count': len(in_file_dups),
            'unique_after_dedup': len(deduped),
            'duplicate_count': len(duplicates),
            'exact_property_matches': exact_property_matches,
            'needs_property_review': needs_property_review,
            'ready_count': ready_count,
        },
    }


# ---- Resolution application (post-user-choice) -------------------------------

def _apply_property_resolutions(
    properties_resolution: List[Dict[str, Any]],
    user_actions: Dict[str, str],
    all_properties_by_id: Dict[int, Property],
) -> Tuple[Dict[str, Optional[Property]], List[Property], List[str]]:
    """Build a name -> Property (or None) map per the user's choices.

    Returns (name_to_property, newly_created_properties, warnings).

    user_actions: dict keyed by property_name from the resolution list. Values:
        - "create_new"
        - "use_existing:<property_id>"
        - "skip"
        (For 'exact' entries, no action key is needed — auto-applied.)
    """
    name_to_property: Dict[str, Optional[Property]] = {}
    newly_created: List[Property] = []
    warnings: List[str] = []

    for p in properties_resolution:
        name = p['name']
        status = p['status']

        if status == 'exact':
            existing = all_properties_by_id.get(p['matched_property_id'])
            name_to_property[name] = existing
            continue

        action = (user_actions.get(name) or '').strip()
        if not action:
            # Default for fuzzy = use the suggestion; default for missing = skip.
            action = 'use_existing:%d' % p['matched_property_id'] if status == 'fuzzy' else 'skip'

        if action == 'create_new':
            address = (p.get('sample_address') or '').strip() or '(set after import)'
            new_prop = Property(
                name=name,
                address_line1=address,
                type='single_family',
            )
            tenant_add(new_prop)
            try:
                get_tenant_session().flush()
            except Exception as exc:
                warnings.append(f"Could not create property '{name}': {exc}")
                name_to_property[name] = None
                continue
            newly_created.append(new_prop)
            name_to_property[name] = new_prop
            if not p.get('sample_address'):
                warnings.append(
                    f"Created property '{name}' with placeholder address — please update it on the property page."
                )
        elif action.startswith('use_existing:'):
            try:
                prop_id = int(action.split(':', 1)[1])
            except (ValueError, IndexError):
                prop_id = None
            chosen = all_properties_by_id.get(prop_id) if prop_id else None
            name_to_property[name] = chosen
            if chosen is None:
                warnings.append(f"Could not resolve selection for property '{name}' — left unassigned.")
        else:
            # 'skip' / unknown -> leave unassigned
            name_to_property[name] = None

    return name_to_property, newly_created, warnings


def _resolve_unit_for_row(
    property_obj: Optional[Property],
    unit_label: str,
    is_property_newly_created: bool,
    auto_create_units: bool,
    units_cache: Dict[int, Dict[str, PropertyUnit]],
    row_idx: int,
) -> Tuple[Optional[PropertyUnit], List[str]]:
    """Resolve a single row's unit. Returns (unit, warnings).

    Auto-creates units for brand-new properties unconditionally. For existing
    properties, only auto-creates if the auto_create_units checkbox was on.
    """
    warnings: List[str] = []
    label = (unit_label or '').strip()

    if not label:
        return None, warnings

    if property_obj is None:
        warnings.append(
            f"Row {row_idx}: unit '{label}' could not be assigned (no resolved property) — left unassigned."
        )
        return None, warnings

    cache = units_cache.setdefault(property_obj.id, {})
    if not cache:
        # Lazy-load units the first time we see this property
        for u in tenant_query(PropertyUnit).filter(PropertyUnit.property_id == property_obj.id).all():
            key = _normalize(u.label)
            if key:
                cache[key] = u

    norm = _normalize(label)
    existing = cache.get(norm)
    if existing is not None:
        return existing, warnings

    # Try fuzzy
    fuzzy_unit, is_fuzzy = _fuzzy_find_unit(label, list(cache.values()))
    if fuzzy_unit is not None:
        warnings.append(
            f"Row {row_idx}: unit '{label}' fuzzy-matched to '{fuzzy_unit.label}' on '{property_obj.name}'."
        )
        return fuzzy_unit, warnings

    # No match — auto-create or warn
    should_create = is_property_newly_created or auto_create_units
    if should_create:
        new_unit = PropertyUnit(property_id=property_obj.id, label=label)
        tenant_add(new_unit)
        try:
            get_tenant_session().flush()
        except Exception as exc:
            warnings.append(f"Row {row_idx}: could not create unit '{label}' on '{property_obj.name}': {exc}")
            return None, warnings
        cache[norm] = new_unit
        if not is_property_newly_created:
            warnings.append(
                f"Row {row_idx}: unit '{label}' created on existing property '{property_obj.name}'."
            )
        return new_unit, warnings

    warnings.append(
        f"Row {row_idx}: unit '{label}' did not match any unit on '{property_obj.name}' — left unassigned."
    )
    return None, warnings


def _extract_item_data(row: Dict[str, str], mapping: Dict[str, str], fields: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """Convert a single CSV row into a dict of import-controlled field values."""
    item_data: Dict[str, Any] = {}
    label_col = mapping.get('label')
    label = (row.get(label_col, '') or '').strip() if label_col else ''
    item_data['label'] = label

    for field_name, config in fields.items():
        if field_name == 'label':
            continue
        column = mapping.get(field_name)
        if not column:
            if 'default' in config:
                item_data[field_name] = config['default']
            continue
        value = (row.get(column, '') or '').strip()
        if not value:
            if 'default' in config:
                item_data[field_name] = config['default']
            continue
        if config.get('type') == 'int':
            try:
                item_data[field_name] = int(value)
            except (ValueError, TypeError):
                item_data[field_name] = config.get('default', 0)
        else:
            item_data[field_name] = value
    return item_data


def _parse_resolution_form(
    form,
    properties_resolution: List[Dict[str, Any]],
    duplicates: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """Pull user choices out of the resolve form."""
    property_actions: Dict[str, str] = {}
    for idx, p in enumerate(properties_resolution):
        if p['status'] == 'exact':
            continue
        action = (form.get(f'prop_action_{idx}') or '').strip()
        if action == 'use_existing':
            chosen_id = (form.get(f'prop_existing_{idx}') or '').strip()
            if chosen_id:
                action = f'use_existing:{chosen_id}'
        property_actions[p['name']] = action

    duplicate_actions: Dict[int, str] = {}
    for idx, d in enumerate(duplicates):
        action = (form.get(f'dup_action_{idx}') or 'skip').strip()
        duplicate_actions[d['row_idx']] = action

    auto_create_units = (form.get('auto_create_units') or '').lower() in ('1', 'true', 'on', 'yes')

    return {
        'property_actions': property_actions,
        'duplicate_actions': duplicate_actions,
        'auto_create_units': auto_create_units,
    }


def _run_resolved_import(
    *,
    analysis: Dict[str, Any],
    mapping: Dict[str, str],
    fields: Dict[str, Dict[str, Any]],
    item_type_canonical: str,
    user_choices: Dict[str, Any],
    custom_id_factory,                # callable(item_data: dict) -> str
) -> Dict[str, Any]:
    """Execute an import once the user has resolved properties + duplicates.

    Returns a dict with counts and a warnings list.
    """
    rows = analysis['rows']                # [(row_idx, row_dict), ...]
    properties_resolution = analysis['properties']
    duplicates = analysis['duplicates']

    auto_create_units = user_choices['auto_create_units']
    property_actions = user_choices['property_actions']
    duplicate_actions = user_choices['duplicate_actions']

    # Pre-load tenant data
    all_properties = tenant_query(Property).all()
    all_properties_by_id: Dict[int, Property] = {p.id: p for p in all_properties}

    name_to_property, newly_created, prop_warnings = _apply_property_resolutions(
        properties_resolution, property_actions, all_properties_by_id,
    )
    newly_created_ids = {p.id for p in newly_created}

    # Pre-load existing items by id for duplicate updates/replaces
    existing_ids_needed = {d['existing_id'] for d in duplicates}
    existing_items_by_id: Dict[int, Item] = {}
    if existing_ids_needed:
        existing_items_by_id = {
            item.id: item
            for item in tenant_query(Item)
                .filter(Item.id.in_(existing_ids_needed))
                .all()
        }
    duplicate_lookup: Dict[int, Dict[str, Any]] = {d['row_idx']: d for d in duplicates}

    units_cache: Dict[int, Dict[str, PropertyUnit]] = {}

    counts = {
        'created': 0,
        'updated': 0,
        'replaced': 0,
        'skipped_duplicate': 0,
        'skipped_blank_label': 0,
        'failed': 0,
    }
    warnings: List[str] = list(prop_warnings)

    # In-file dedup notice
    in_file_dups = analysis.get('in_file_dups') or []
    if in_file_dups:
        sample = ', '.join(f"row {r} ({lbl})" for r, lbl, _ in in_file_dups[:5])
        more = '' if len(in_file_dups) <= 5 else f" (+{len(in_file_dups)-5} more)"
        warnings.append(
            f"{len(in_file_dups)} duplicate label(s) within the file were skipped: {sample}{more}"
        )

    for idx, row in rows:
        try:
            item_data = _extract_item_data(row, mapping, fields)
            label = item_data.get('label', '').strip()
            if not label:
                counts['skipped_blank_label'] += 1
                warnings.append(f"Row {idx}: blank label — skipped.")
                continue

            property_name = item_data.pop('property_name', None)
            property_unit_label = item_data.pop('property_unit_label', None)

            property_obj = name_to_property.get(property_name) if property_name else None
            is_newly_created_prop = bool(property_obj and property_obj.id in newly_created_ids)

            unit_obj, unit_warnings = _resolve_unit_for_row(
                property_obj=property_obj,
                unit_label=property_unit_label or '',
                is_property_newly_created=is_newly_created_prop,
                auto_create_units=auto_create_units,
                units_cache=units_cache,
                row_idx=idx,
            )
            warnings.extend(unit_warnings)

            dup = duplicate_lookup.get(idx)
            if dup:
                action = duplicate_actions.get(idx, 'skip')
                existing = existing_items_by_id.get(dup['existing_id'])
                if existing is None:
                    # Fell out from under us between analysis and apply.
                    warnings.append(
                        f"Row {idx}: duplicate target missing for '{label}' — skipped."
                    )
                    counts['skipped_duplicate'] += 1
                    continue

                if action == 'skip':
                    counts['skipped_duplicate'] += 1
                    continue

                if action == 'update':
                    # Merge: only set non-empty values from the row.
                    for field_name, value in item_data.items():
                        if value in (None, ''):
                            continue
                        if hasattr(existing, field_name):
                            setattr(existing, field_name, value)
                    if property_obj is not None:
                        existing.property_id = property_obj.id
                    if unit_obj is not None:
                        existing.property_unit_id = unit_obj.id
                    existing.last_action = 'updated'
                    existing.last_action_at = utc_now()
                    existing.last_action_by_id = current_user.id if current_user.is_authenticated else None
                    counts['updated'] += 1
                    continue

                if action == 'replace':
                    # Overwrite mapped columns only. Empty cells in a mapped
                    # column blank out the existing value (this is the
                    # difference from "update"). Unmapped columns stay
                    # untouched — Replace doesn't wipe fields the user
                    # never expressed an opinion about.
                    for field_name, config in fields.items():
                        if field_name in ('property_name', 'property_unit_label'):
                            continue
                        if field_name not in mapping:
                            continue
                        if field_name in item_data:
                            setattr(existing, field_name, item_data[field_name])
                        else:
                            # Mapped column but empty cell: blank out (or
                            # fall back to declared default for fields like
                            # status which is nullable=False).
                            default = config.get('default', None)
                            setattr(existing, field_name, default)
                    if 'property_name' in mapping:
                        existing.property_id = property_obj.id if property_obj else None
                    if 'property_unit_label' in mapping:
                        existing.property_unit_id = unit_obj.id if unit_obj else None
                    existing.last_action = 'updated'
                    existing.last_action_at = utc_now()
                    existing.last_action_by_id = current_user.id if current_user.is_authenticated else None
                    counts['replaced'] += 1
                    continue

                # Unknown action — treat as skip
                counts['skipped_duplicate'] += 1
                continue

            # New item
            new_item = Item(
                type=item_type_canonical,
                custom_id=custom_id_factory(item_data),
                property_id=property_obj.id if property_obj else None,
                property_unit_id=unit_obj.id if unit_obj else None,
                last_action='added',
                last_action_at=utc_now(),
                last_action_by_id=current_user.id if current_user.is_authenticated else None,
                **item_data,
            )
            tenant_add(new_item)
            counts['created'] += 1

        except Exception as exc:
            counts['failed'] += 1
            warnings.append(f"Row {idx}: {exc}")

    return {'counts': counts, 'warnings': warnings}


def _flash_import_summary(counts: Dict[str, int], item_type_plural: str) -> None:
    """Compose the final result flash."""
    parts = []
    if counts['created']:
        parts.append(f"Imported {counts['created']:,} new")
    if counts['updated']:
        parts.append(f"updated {counts['updated']:,}")
    if counts['replaced']:
        parts.append(f"replaced {counts['replaced']:,}")
    if counts['skipped_duplicate']:
        parts.append(f"skipped {counts['skipped_duplicate']:,} duplicate(s)")
    if counts['skipped_blank_label']:
        parts.append(f"skipped {counts['skipped_blank_label']:,} blank-label row(s)")
    if counts['failed']:
        parts.append(f"failed {counts['failed']:,}")

    msg_type = 'success' if (counts['created'] or counts['updated'] or counts['replaced']) else 'warning'
    if parts:
        flash(f"{item_type_plural}: " + ' · '.join(parts) + '.', msg_type)
    else:
        flash(f"No {item_type_plural.lower()} were imported.", 'warning')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def parse_csv_file(file_content: str) -> tuple[List[str], List[Dict[str, str]]]:
    """Parse CSV file and return headers and rows"""
    reader = csv.DictReader(io.StringIO(file_content))
    headers = reader.fieldnames or []
    rows = [row for row in reader]
    return list(headers), rows


def parse_excel_file(file_bytes: bytes) -> tuple[List[str], List[Dict[str, str]]]:
    """Parse Excel file and return headers and rows"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel import")

    # Load workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Get all rows as list
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        return [], []

    # First row is headers
    headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(all_rows[0])]

    # Rest are data rows
    rows = []
    for row_data in all_rows[1:]:
        # Convert row to dictionary, handling None values
        row_dict = {}
        for i, cell_value in enumerate(row_data):
            if i < len(headers):
                # Convert to string, handle None and numeric values
                if cell_value is None:
                    row_dict[headers[i]] = ''
                else:
                    row_dict[headers[i]] = str(cell_value)
        rows.append(row_dict)

    wb.close()
    return headers, rows


KEY_FIELDS = {
    'label': {'required': True, 'name': 'Label'},
    'key_hook_number': {'required': False, 'name': 'Key Hook Number'},
    'keycode': {'required': False, 'name': 'Key Code'},
    'total_copies': {'required': False, 'name': 'Total Copies', 'type': 'int'},
    'location': {'required': False, 'name': 'Key Box Location'},
    'address': {'required': False, 'name': 'Address'},
    'status': {'required': False, 'name': 'Status', 'default': 'available'},
    'assigned_to': {'required': False, 'name': 'Assigned To'},
    'property_name': {'required': False, 'name': 'Property Name'},
    'property_unit_label': {'required': False, 'name': 'Property Unit Label'},
}

LOCKBOX_FIELDS = {
    'label': {'required': True, 'name': 'Label'},
    'code_current': {'required': False, 'name': 'Current Code'},
    'code_previous': {'required': False, 'name': 'Previous Code'},
    'supra_id': {'required': False, 'name': 'Supra ID'},
    'location': {'required': False, 'name': 'Location'},
    'address': {'required': False, 'name': 'Address'},
    'status': {'required': False, 'name': 'Status', 'default': 'available'},
    'assigned_to': {'required': False, 'name': 'Assigned To'},
    'property_name': {'required': False, 'name': 'Property Name'},
    'property_unit_label': {'required': False, 'name': 'Property Unit Label'},
}

SIGN_FIELDS = {
    'label': {'required': True, 'name': 'Label'},
    'sign_subtype': {'required': False, 'name': 'Sign Type (Piece/Assembled Unit)'},
    'piece_type': {'required': False, 'name': 'Piece Type'},
    'rider_text': {'required': False, 'name': 'Rider Text'},
    'material': {'required': False, 'name': 'Material'},
    'condition': {'required': False, 'name': 'Condition'},
    'location': {'required': False, 'name': 'Storage Location'},
    'address': {'required': False, 'name': 'Property Address'},
    'status': {'required': False, 'name': 'Status', 'default': 'available'},
    'assigned_to': {'required': False, 'name': 'Assigned To'},
    'property_name': {'required': False, 'name': 'Property Name'},
    'property_unit_label': {'required': False, 'name': 'Property Unit Label'},
}


@inventory_bp.route("/keys/import", methods=["GET", "POST"])
@login_required
@tenant_required
def import_keys():
    """Import keys from CSV/Excel file"""
    if request.method == "POST":
        # Check if file was uploaded
        if 'file' not in request.files:
            flash("No file uploaded", "error")
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash("No file selected", "error")
            return redirect(request.url)

        if not allowed_file(file.filename):
            flash("Invalid file type. Please upload a CSV or Excel file.", "error")
            return redirect(request.url)

        try:
            # Parse file based on extension
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower()

            if file_ext == 'csv':
                file_content = file.read().decode('utf-8')
                headers, rows = parse_csv_file(file_content)
            else:  # Excel
                file_bytes = file.read()
                headers, rows = parse_excel_file(file_bytes)

            if not headers or not rows:
                flash("File is empty or could not be parsed", "error")
                return redirect(request.url)

            # Cap the rows we keep in the session. Past MAX_IMPORT_ROWS we'd
            # blow session size for cookie-based stores.
            total_rows = len(rows)
            kept_rows = rows[:MAX_IMPORT_ROWS]
            if total_rows > MAX_IMPORT_ROWS:
                flash(
                    f"File has {total_rows:,} rows; only the first {MAX_IMPORT_ROWS:,} "
                    "will be imported. Split larger files and re-upload to import the rest.",
                    "warning",
                )

            session['import_data'] = {
                'headers': headers,
                'rows': kept_rows,
                'total_rows': total_rows,
                'file_type': 'keys',
            }

            return redirect(url_for('inventory.import_keys_map'))

        except Exception as e:
            flash(f"Error processing file: {str(e)}", "error")
            return redirect(request.url)

    return render_template("import_upload.html", item_type="Keys")


@inventory_bp.route("/keys/import/map", methods=["GET", "POST"])
@login_required
@tenant_required
def import_keys_map():
    """Map columns from uploaded file to database fields"""
    import_data = session.get('import_data')
    if not import_data:
        flash("No import data found. Please upload a file first.", "error")
        return redirect(url_for('inventory.import_keys'))

    if request.method == "POST":
        # Get column mapping from form
        mapping = {}
        for field_name in KEY_FIELDS.keys():
            column = request.form.get(f'map_{field_name}')
            if column and column != '':
                mapping[field_name] = column

        # Validate required fields
        if 'label' not in mapping:
            flash("Label field is required", "error")
            return redirect(request.url)

        # Store mapping in session
        session['import_mapping'] = mapping
        return redirect(url_for('inventory.import_keys_process'))

    return render_template(
        "import_map.html",
        import_data=import_data,
        fields=KEY_FIELDS,
        item_type="Keys"
    )


@inventory_bp.route("/keys/import/process", methods=["GET", "POST"])
@login_required
@tenant_required
def import_keys_process():
    """Resolve duplicates + properties, then import.

    GET: render the resolve wizard with classified rows, duplicates, and
         properties needing user attention.
    POST: re-analyze (cheap), apply user decisions, run the import, redirect
          back to the list with a summary flash.
    """
    import_data = session.get('import_data')
    mapping = session.get('import_mapping')

    if not import_data or not mapping:
        flash("Import session expired. Please start over.", "error")
        return redirect(url_for('inventory.import_keys'))

    rows = import_data['rows']

    if request.method == "POST":
        analysis = _analyze_import(rows, mapping, 'Key')
        user_choices = _parse_resolution_form(request.form, analysis['properties'], analysis['duplicates'])

        try:
            result = _run_resolved_import(
                analysis=analysis,
                mapping=mapping,
                fields=KEY_FIELDS,
                item_type_canonical='Key',
                user_choices=user_choices,
                custom_id_factory=lambda item_data: Item.generate_custom_id('Key'),
            )
            tenant_commit()
        except Exception as exc:
            tenant_rollback()
            flash(f"Database error: {exc}", "error")
            return redirect(url_for('inventory.import_keys'))

        _flash_import_summary(result['counts'], 'Keys')
        _flash_warning_summary(result['warnings'])

        session.pop('import_data', None)
        session.pop('import_mapping', None)
        return redirect(url_for('inventory.list_keys'))

    # GET — render the resolve wizard
    analysis = _analyze_import(rows, mapping, 'Key')
    return render_template(
        "import_resolve.html",
        analysis=analysis,
        mapping=mapping,
        fields=KEY_FIELDS,
        item_type="Keys",
        item_type_singular="Key",
        process_url=url_for('inventory.import_keys_process'),
        cancel_url=url_for('inventory.import_keys'),
        back_url=url_for('inventory.import_keys_map'),
    )


@inventory_bp.route("/lockboxes/import", methods=["GET", "POST"])
@login_required
@tenant_required
def import_lockboxes():
    """Import lockboxes from CSV/Excel file"""
    if request.method == "POST":
        # Check if file was uploaded
        if 'file' not in request.files:
            flash("No file uploaded", "error")
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash("No file selected", "error")
            return redirect(request.url)

        if not allowed_file(file.filename):
            flash("Invalid file type. Please upload a CSV or Excel file.", "error")
            return redirect(request.url)

        try:
            # Parse file based on extension
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower()

            if file_ext == 'csv':
                file_content = file.read().decode('utf-8')
                headers, rows = parse_csv_file(file_content)
            else:  # Excel
                file_bytes = file.read()
                headers, rows = parse_excel_file(file_bytes)

            if not headers or not rows:
                flash("File is empty or could not be parsed", "error")
                return redirect(request.url)

            total_rows = len(rows)
            kept_rows = rows[:MAX_IMPORT_ROWS]
            if total_rows > MAX_IMPORT_ROWS:
                flash(
                    f"File has {total_rows:,} rows; only the first {MAX_IMPORT_ROWS:,} "
                    "will be imported. Split larger files and re-upload to import the rest.",
                    "warning",
                )

            session['import_data'] = {
                'headers': headers,
                'rows': kept_rows,
                'total_rows': total_rows,
                'file_type': 'lockboxes',
            }

            return redirect(url_for('inventory.import_lockboxes_map'))

        except Exception as e:
            flash(f"Error processing file: {str(e)}", "error")
            return redirect(request.url)

    return render_template("import_upload.html", item_type="Lockboxes")


@inventory_bp.route("/lockboxes/import/map", methods=["GET", "POST"])
@login_required
@tenant_required
def import_lockboxes_map():
    """Map columns from uploaded file to database fields"""
    import_data = session.get('import_data')
    if not import_data:
        flash("No import data found. Please upload a file first.", "error")
        return redirect(url_for('inventory.import_lockboxes'))

    if request.method == "POST":
        # Get column mapping from form
        mapping = {}
        for field_name in LOCKBOX_FIELDS.keys():
            column = request.form.get(f'map_{field_name}')
            if column and column != '':
                mapping[field_name] = column

        # Validate required fields
        if 'label' not in mapping:
            flash("Label field is required", "error")
            return redirect(request.url)

        # Store mapping in session
        session['import_mapping'] = mapping
        return redirect(url_for('inventory.import_lockboxes_process'))

    return render_template(
        "import_map.html",
        import_data=import_data,
        fields=LOCKBOX_FIELDS,
        item_type="Lockboxes"
    )


@inventory_bp.route("/lockboxes/import/process", methods=["GET", "POST"])
@login_required
@tenant_required
def import_lockboxes_process():
    """Resolve duplicates + properties, then import lockboxes."""
    import_data = session.get('import_data')
    mapping = session.get('import_mapping')

    if not import_data or not mapping:
        flash("Import session expired. Please start over.", "error")
        return redirect(url_for('inventory.import_lockboxes'))

    rows = import_data['rows']

    if request.method == "POST":
        analysis = _analyze_import(rows, mapping, 'Lockbox')
        user_choices = _parse_resolution_form(request.form, analysis['properties'], analysis['duplicates'])

        try:
            result = _run_resolved_import(
                analysis=analysis,
                mapping=mapping,
                fields=LOCKBOX_FIELDS,
                item_type_canonical='Lockbox',
                user_choices=user_choices,
                custom_id_factory=lambda item_data: Item.generate_custom_id('Lockbox'),
            )
            tenant_commit()
        except Exception as exc:
            tenant_rollback()
            flash(f"Database error: {exc}", "error")
            return redirect(url_for('inventory.import_lockboxes'))

        _flash_import_summary(result['counts'], 'Lockboxes')
        _flash_warning_summary(result['warnings'])

        session.pop('import_data', None)
        session.pop('import_mapping', None)
        return redirect(url_for('inventory.list_lockboxes'))

    analysis = _analyze_import(rows, mapping, 'Lockbox')
    return render_template(
        "import_resolve.html",
        analysis=analysis,
        mapping=mapping,
        fields=LOCKBOX_FIELDS,
        item_type="Lockboxes",
        item_type_singular="Lockbox",
        process_url=url_for('inventory.import_lockboxes_process'),
        cancel_url=url_for('inventory.import_lockboxes'),
        back_url=url_for('inventory.import_lockboxes_map'),
    )


@inventory_bp.route("/signs/import", methods=["GET", "POST"])
@login_required
@tenant_required
def import_signs():
    """Import signs from CSV/Excel file"""
    if request.method == "POST":
        # Check if file was uploaded
        if 'file' not in request.files:
            flash("No file uploaded", "error")
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash("No file selected", "error")
            return redirect(request.url)

        if not allowed_file(file.filename):
            flash("Invalid file type. Please upload a CSV or Excel file.", "error")
            return redirect(request.url)

        try:
            # Parse file based on extension
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower()

            if file_ext == 'csv':
                file_content = file.read().decode('utf-8')
                headers, rows = parse_csv_file(file_content)
            else:  # Excel
                file_bytes = file.read()
                headers, rows = parse_excel_file(file_bytes)

            if not headers or not rows:
                flash("File is empty or could not be parsed", "error")
                return redirect(request.url)

            total_rows = len(rows)
            kept_rows = rows[:MAX_IMPORT_ROWS]
            if total_rows > MAX_IMPORT_ROWS:
                flash(
                    f"File has {total_rows:,} rows; only the first {MAX_IMPORT_ROWS:,} "
                    "will be imported. Split larger files and re-upload to import the rest.",
                    "warning",
                )

            session['import_data'] = {
                'headers': headers,
                'rows': kept_rows,
                'total_rows': total_rows,
                'file_type': 'signs',
            }

            return redirect(url_for('inventory.import_signs_map'))

        except Exception as e:
            flash(f"Error parsing file: {str(e)}", "error")
            return redirect(request.url)

    return render_template("import_upload.html", item_type="Signs")


@inventory_bp.route("/signs/import/map", methods=["GET", "POST"])
@login_required
@tenant_required
def import_signs_map():
    """Map CSV/Excel columns to sign fields"""
    import_data = session.get('import_data')
    if not import_data:
        flash("No import data found. Please upload a file first.", "error")
        return redirect(url_for('inventory.import_signs'))

    if request.method == "POST":
        # Get column mapping from form
        mapping = {}
        for field_name in SIGN_FIELDS.keys():
            column = request.form.get(f'map_{field_name}')
            if column and column != '':
                mapping[field_name] = column

        # Validate required fields
        for field_name, config in SIGN_FIELDS.items():
            if config['required'] and field_name not in mapping:
                flash(f"Please map the required field: {config['name']}", "error")
                return redirect(url_for('inventory.import_signs_map'))

        session['import_mapping'] = mapping
        return redirect(url_for('inventory.import_signs_process'))

    return render_template(
        "import_map.html",
        import_data=import_data,
        fields=SIGN_FIELDS,
        item_type="Signs"
    )


@inventory_bp.route("/signs/import/process", methods=["GET", "POST"])
@login_required
@tenant_required
def import_signs_process():
    """Resolve duplicates + properties, then import signs."""
    import_data = session.get('import_data')
    mapping = session.get('import_mapping')

    if not import_data or not mapping:
        flash("Missing import data. Please start the import process again.", "error")
        return redirect(url_for('inventory.import_signs'))

    rows = import_data['rows']

    if request.method == "POST":
        analysis = _analyze_import(rows, mapping, 'Sign')
        user_choices = _parse_resolution_form(request.form, analysis['properties'], analysis['duplicates'])

        try:
            result = _run_resolved_import(
                analysis=analysis,
                mapping=mapping,
                fields=SIGN_FIELDS,
                item_type_canonical='Sign',
                user_choices=user_choices,
                custom_id_factory=lambda item_data: Item.generate_custom_id('Sign', item_data.get('sign_subtype')),
            )
            tenant_commit()
        except Exception as exc:
            tenant_rollback()
            flash(f"Database error: {exc}", "error")
            return redirect(url_for('inventory.import_signs'))

        _flash_import_summary(result['counts'], 'Signs')
        _flash_warning_summary(result['warnings'])

        session.pop('import_data', None)
        session.pop('import_mapping', None)
        return redirect(url_for('inventory.list_signs'))

    analysis = _analyze_import(rows, mapping, 'Sign')
    return render_template(
        "import_resolve.html",
        analysis=analysis,
        mapping=mapping,
        fields=SIGN_FIELDS,
        item_type="Signs",
        item_type_singular="Sign",
        process_url=url_for('inventory.import_signs_process'),
        cancel_url=url_for('inventory.import_signs'),
        back_url=url_for('inventory.import_signs_map'),
    )

# exports/views.py
from flask import Blueprint, Response, request, render_template, flash, redirect, url_for, jsonify
from flask_login import login_required, current_user
from io import BytesIO, StringIO
import csv
from datetime import datetime
from typing import List, Dict, Any

from utilities.tenant_helpers import tenant_query, tenant_add, tenant_commit, tenant_rollback, get_tenant_session
from middleware.tenant_middleware import tenant_required
from utilities.database import db, Item, ItemCheckout, Contact, Property, PropertyUnit, User

exports_bp = Blueprint("exports", __name__, template_folder="../templates")


def generate_csv(data: List[Dict[str, Any]], filename: str) -> Response:
    """Generate CSV file from data"""
    if not data:
        return Response("No data to export", status=400)

    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)

    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def generate_excel(data: List[Dict[str, Any]], filename: str) -> Response:
    """Generate Excel file from data"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as e:
        error_msg = f"Excel export not available - openpyxl import failed: {str(e)}"
        print(error_msg)  # Log to console
        return Response(error_msg, status=500)

    if not data:
        return Response("No data to export", status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    # Write headers
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    try:
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        error_msg = f'Excel generation error: {str(e)}'
        print(error_msg)
        return Response(error_msg, status=500)


def generate_pdf(data: List[Dict[str, Any]], filename: str, title: str = "Report") -> Response:
    """Generate PDF file from data with KBM styling and improved formatting"""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
    except ImportError as e:
        error_msg = f"PDF export not available - reportlab import failed: {str(e)}"
        print(error_msg)  # Log to console
        return Response(error_msg, status=500)

    if not data:
        return Response("No data to export", status=400)

    output = BytesIO()

    # Use smaller margins for more space
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        topMargin=0.5*inch,
        bottomMargin=0.5*inch,
        leftMargin=0.5*inch,
        rightMargin=0.5*inch
    )
    elements = []

    # Custom KBM styles
    styles = getSampleStyleSheet()

    # KBM brand header style
    kbm_header = ParagraphStyle(
        'KBMHeader',
        parent=styles['Normal'],
        fontSize=20,
        textColor=colors.HexColor('#e53935'),
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        spaceAfter=4
    )

    # Title style
    title_style = ParagraphStyle(
        'KBMTitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#1f2937'),
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        spaceAfter=8
    )

    # Metadata style
    meta_style = ParagraphStyle(
        'KBMMeta',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#6b7280'),
        alignment=TA_CENTER,
        spaceAfter=12
    )

    # Header with KBM branding
    brand_para = Paragraph('KBM <font color="#e53935">2.0</font>', kbm_header)
    elements.append(brand_para)

    subtitle_para = Paragraph('Key & Lockbox Management System', meta_style)
    elements.append(subtitle_para)

    # Title
    title_para = Paragraph(title, title_style)
    elements.append(title_para)

    # Export metadata
    export_date = datetime.now().strftime('%B %d, %Y at %I:%M %p')
    meta_para = Paragraph(f'Exported on {export_date} | Total Records: {len(data)}', meta_style)
    elements.append(meta_para)

    elements.append(Spacer(1, 0.15 * inch))

    # Prepare table data with text wrapping support
    headers = list(data[0].keys())
    num_columns = len(headers)

    # Calculate available width
    page_width = landscape(letter)[0]
    available_width = page_width - (doc.leftMargin + doc.rightMargin)

    # Dynamically adjust font size based on column count
    if num_columns <= 6:
        data_font_size = 9
        header_font_size = 10
    elif num_columns <= 9:
        data_font_size = 8
        header_font_size = 9
    else:
        data_font_size = 7
        header_font_size = 8

    # Cell text style for wrapping
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontSize=data_font_size,
        leading=data_font_size + 2,
        textColor=colors.HexColor('#1f2937'),
        fontName='Helvetica'
    )

    header_cell_style = ParagraphStyle(
        'HeaderCellText',
        parent=styles['Normal'],
        fontSize=header_font_size,
        leading=header_font_size + 2,
        textColor=colors.white,
        fontName='Helvetica-Bold'
    )

    # Build table data with Paragraph objects for text wrapping
    table_data = []

    # Header row with Paragraphs
    header_row = [Paragraph(str(h), header_cell_style) for h in headers]
    table_data.append(header_row)

    # Data rows with Paragraphs
    for row in data:
        row_data = []
        for h in headers:
            value = row.get(h, '')
            if value in [None, '']:
                value = '-'
            else:
                value = str(value)
            row_data.append(Paragraph(value, cell_style))
        table_data.append(row_data)

    # Calculate proportional column widths
    # Give more space to columns that typically have longer text
    col_widths = []
    priority_columns = ['Label', 'Address', 'Location', 'Property', 'Assigned To']

    base_width = available_width / num_columns
    for header in headers:
        if any(priority in header for priority in priority_columns):
            col_widths.append(base_width * 1.3)  # 30% wider for important columns
        else:
            col_widths.append(base_width * 0.8)  # 20% narrower for short columns

    # Normalize widths to fit available space
    total_width = sum(col_widths)
    col_widths = [w * (available_width / total_width) for w in col_widths]

    # KBM accent colors
    kbm_red = colors.HexColor('#e53935')
    light_gray = colors.HexColor('#f5f7fb')
    border_gray = colors.HexColor('#e0e0e0')
    text_dark = colors.HexColor('#1f2937')

    # Create table with calculated column widths
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header row styling
        ('BACKGROUND', (0, 0), (-1, 0), kbm_red),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),

        # Data rows styling
        ('BACKGROUND', (0, 1), (-1, -1), light_gray),
        ('TEXTCOLOR', (0, 1), (-1, -1), text_dark),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), data_font_size),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),

        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, border_gray),
        ('BOX', (0, 0), (-1, -1), 1.5, kbm_red),

        # Alternating row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [light_gray, colors.white]),
    ]))

    elements.append(table)

    try:
        doc.build(elements)
        output.seek(0)

        response = Response(output.getvalue(), mimetype='application/pdf')
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        error_msg = f'PDF generation error: {str(e)}'
        print(error_msg)
        return Response(error_msg, status=500)




@exports_bp.route("/preview/<item_type>", methods=["GET"])
@login_required
@tenant_required
def preview_export(item_type):
    """Return preview data for export (first 20 rows)"""
    valid_types = ['keys', 'lockboxes', 'signs', 'properties', 'contacts']
    if item_type not in valid_types:
        return jsonify({'error': 'Invalid item type'}), 400

    preview_data = {
        'item_type': item_type,
        'total_count': 0,
        'preview_count': 0,
        'headers': [],
        'rows': []
    }

    # Get preview data based on item type
    if item_type == 'keys':
        items = tenant_query(Item).filter_by(type='Key').order_by(Item.id.desc()).limit(20).all()
        preview_data['total_count'] = tenant_query(Item).filter_by(type='Key').count()
    elif item_type == 'lockboxes':
        items = tenant_query(Item).filter_by(type='Lockbox').order_by(Item.id.desc()).limit(20).all()
        preview_data['total_count'] = tenant_query(Item).filter_by(type='Lockbox').count()
    elif item_type == 'signs':
        items = tenant_query(Item).filter_by(type='Sign').order_by(Item.id.desc()).limit(20).all()
        preview_data['total_count'] = tenant_query(Item).filter_by(type='Sign').count()
    elif item_type == 'properties':
        properties = tenant_query(Property).order_by(Property.id.desc()).limit(20).all()
        preview_data['total_count'] = tenant_query(Property).count()
        preview_data['preview_count'] = len(properties)
        if properties:
            data_dicts = [p.to_dict() for p in properties]
            preview_data['headers'] = list(data_dicts[0].keys()) if data_dicts else []
            preview_data['rows'] = data_dicts
        return jsonify(preview_data)
    elif item_type == 'contacts':
        contacts = tenant_query(Contact).order_by(Contact.id.desc()).limit(20).all()
        preview_data['total_count'] = tenant_query(Contact).count()
        preview_data['preview_count'] = len(contacts)
        if contacts:
            data_dicts = [c.to_dict() for c in contacts]
            preview_data['headers'] = list(data_dicts[0].keys()) if data_dicts else []
            preview_data['rows'] = data_dicts
        return jsonify(preview_data)
    else:
        items = []

    # For item types (keys, lockboxes, signs)
    preview_data['preview_count'] = len(items)
    if items:
        data_dicts = [item.to_dict() for item in items]
        preview_data['headers'] = list(data_dicts[0].keys()) if data_dicts else []
        preview_data['rows'] = data_dicts

    return jsonify(preview_data)

@exports_bp.route("/items/<item_type>", methods=["GET"])
@login_required
@tenant_required
def export_items(item_type: str):
    """Export items (keys, lockboxes, signs) to CSV/Excel/PDF"""
    format_type = request.args.get("format", "csv").lower()

    # Validate item type
    valid_types = ["key", "lockbox", "sign"]
    item_type_singular = item_type.rstrip('s').lower()
    if item_type_singular not in valid_types:
        flash("Invalid item type", "error")
        return redirect(url_for("main.home"))

    # Capitalize for database query
    item_type_db = item_type_singular.capitalize()

    # Get items
    items = tenant_query(Item).filter_by(type=item_type_db).order_by(Item.label.asc()).all()

    # Prepare data based on item type
    data = []
    for item in items:
        row = {
            "ID": item.custom_id or "",
            "Label": item.label or "",
            "Status": item.status or "",
            "Location": item.location or "",
            "Address": item.address or "",
        }

        if item_type_db == "Key":
            row.update({
                "Key Hook #": item.key_hook_number or "",
                "Key Code": item.keycode or "",
                "Total Copies": item.total_copies or 0,
                "Available": (item.total_copies or 0) - (item.copies_checked_out or 0),
                "Assigned To": item.assigned_to or "",
            })
        elif item_type_db == "Lockbox":
            row.update({
                "Current Code": item.code_current or "",
                "Previous Code": item.code_previous or "",
            })
        elif item_type_db == "Sign":
            row.update({
                "Subtype": item.sign_subtype or "",
                "Material": item.material or "",
                "Condition": item.condition or "",
            })

        if item.property:
            row["Property"] = item.property.name
        if item.property_unit:
            row["Unit"] = item.property_unit.label

        data.append(row)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if format_type == "excel":
        filename = f"{item_type}_{timestamp}.xlsx"
        return generate_excel(data, filename)
    elif format_type == "pdf":
        filename = f"{item_type}_{timestamp}.pdf"
        title = f"{item_type.title()} Inventory Report"
        return generate_pdf(data, filename, title)
    else:  # csv
        filename = f"{item_type}_{timestamp}.csv"
        return generate_csv(data, filename)


@exports_bp.route("/reports/low-keys", methods=["GET"])
@login_required
@tenant_required
def export_low_keys():
    """Export low key inventory report"""
    format_type = request.args.get("format", "csv").lower()

    keys = tenant_query(Item).filter(
        Item.type == "Key",
        Item.total_copies < 4
    ).order_by(Item.total_copies.asc(), Item.label.asc()).all()

    data = []
    for key in keys:
        data.append({
            "ID": key.custom_id or "",
            "Label": key.label or "",
            "Address": key.address or "",
            "Total Copies": key.total_copies or 0,
            "Available": (key.total_copies or 0) - (key.copies_checked_out or 0),
            "Status": key.status or "",
        })

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if format_type == "excel":
        return generate_excel(data, f"low_keys_{timestamp}.xlsx")
    elif format_type == "pdf":
        return generate_pdf(data, f"low_keys_{timestamp}.pdf", "Low Key Inventory Report")
    else:
        return generate_csv(data, f"low_keys_{timestamp}.csv")


@exports_bp.route("/reports/checked-out-keys", methods=["GET"])
@login_required
@tenant_required
def export_checked_out_keys():
    """Export checked out keys report"""
    format_type = request.args.get("format", "csv").lower()

    keys = tenant_query(Item).filter(
        Item.type == "Key",
        Item.copies_checked_out > 0
    ).order_by(Item.label.asc()).all()

    data = []
    for key in keys:
        data.append({
            "ID": key.custom_id or "",
            "Label": key.label or "",
            "Address": key.address or "",
            "Total Copies": key.total_copies or 0,
            "Checked Out": key.copies_checked_out or 0,
            "Available": (key.total_copies or 0) - (key.copies_checked_out or 0),
            "Assigned To": key.assigned_to or "",
            "Status": key.status or "",
        })

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if format_type == "excel":
        return generate_excel(data, f"checked_out_keys_{timestamp}.xlsx")
    elif format_type == "pdf":
        return generate_pdf(data, f"checked_out_keys_{timestamp}.pdf", "Checked Out Keys Report")
    else:
        return generate_csv(data, f"checked_out_keys_{timestamp}.csv")


@exports_bp.route("/reports/overdue-returns", methods=["GET"])
@login_required
@tenant_required
def export_overdue_returns():
    """Export overdue returns report"""
    format_type = request.args.get("format", "csv").lower()
    
    from datetime import date
    today = date.today()
    
    checkouts = tenant_query(ItemCheckout).filter(
        ItemCheckout.is_active == True,
        ItemCheckout.expected_return_date < today
    ).order_by(ItemCheckout.expected_return_date.asc()).all()
    
    data = []
    for checkout in checkouts:
        item = checkout.item
        data.append({
            'Checked Out To': checkout.checked_out_to,
            'Item Type': item.type if item else 'N/A',
            'Item Label': item.label if item else 'N/A',
            'Item ID': item.custom_id if item else 'N/A',
            'Address': checkout.address or (item.address if item else ''),
            'Checkout Date': checkout.checked_out_at.strftime('%Y-%m-%d %I:%M %p') if checkout.checked_out_at else '',
            'Expected Return': checkout.expected_return_date.strftime('%Y-%m-%d') if checkout.expected_return_date else '',
            'Days Overdue': (today - checkout.expected_return_date).days if checkout.expected_return_date else 0,
            'Assignment Type': checkout.assignment_type or 'checkout',
        })
    
    if format_type == 'csv':
        return generate_csv(data, 'overdue_returns.csv')
    elif format_type == 'excel':
        return generate_excel(data, 'overdue_returns.xlsx')
    elif format_type == 'pdf':
        return generate_pdf(data, 'overdue_returns.pdf', 'Overdue Returns Report')
    else:
        return Response('Invalid format', status=400)


@exports_bp.route("/reports/upcoming-returns", methods=["GET"])
@login_required
@tenant_required
def export_upcoming_returns():
    """Export upcoming returns report (next 30 days)"""
    format_type = request.args.get("format", "csv").lower()
    
    from datetime import date, timedelta
    today = date.today()
    future = today + timedelta(days=30)
    
    checkouts = tenant_query(ItemCheckout).filter(
        ItemCheckout.is_active == True,
        ItemCheckout.expected_return_date.between(today, future)
    ).order_by(ItemCheckout.expected_return_date.asc()).all()
    
    data = []
    for checkout in checkouts:
        item = checkout.item
        data.append({
            'Checked Out To': checkout.checked_out_to,
            'Item Type': item.type if item else 'N/A',
            'Item Label': item.label if item else 'N/A',
            'Item ID': item.custom_id if item else 'N/A',
            'Address': checkout.address or (item.address if item else ''),
            'Checkout Date': checkout.checked_out_at.strftime('%Y-%m-%d %I:%M %p') if checkout.checked_out_at else '',
            'Expected Return': checkout.expected_return_date.strftime('%Y-%m-%d') if checkout.expected_return_date else '',
            'Days Until Return': (checkout.expected_return_date - today).days if checkout.expected_return_date else 0,
            'Assignment Type': checkout.assignment_type or 'checkout',
        })
    
    if format_type == 'csv':
        return generate_csv(data, 'upcoming_returns.csv')
    elif format_type == 'excel':
        return generate_excel(data, 'upcoming_returns.xlsx')
    elif format_type == 'pdf':
        return generate_pdf(data, 'upcoming_returns.pdf', 'Upcoming Returns Report')
    else:
        return Response('Invalid format', status=400)


@exports_bp.route("/reports/long-term-checkouts", methods=["GET"])
@login_required
@tenant_required
def export_long_term_checkouts():
    """Export long-term checkouts report (> 30 days)"""
    format_type = request.args.get("format", "csv").lower()
    
    from datetime import datetime, timedelta
    threshold = datetime.now() - timedelta(days=30)
    
    checkouts = tenant_query(ItemCheckout).filter(
        ItemCheckout.is_active == True,
        ItemCheckout.checked_out_at < threshold
    ).order_by(ItemCheckout.checked_out_at.asc()).all()
    
    data = []
    for checkout in checkouts:
        item = checkout.item
        days_out = (datetime.now() - checkout.checked_out_at).days if checkout.checked_out_at else 0
        data.append({
            'Checked Out To': checkout.checked_out_to,
            'Item Type': item.type if item else 'N/A',
            'Item Label': item.label if item else 'N/A',
            'Item ID': item.custom_id if item else 'N/A',
            'Address': checkout.address or (item.address if item else ''),
            'Checkout Date': checkout.checked_out_at.strftime('%Y-%m-%d %I:%M %p') if checkout.checked_out_at else '',
            'Days Checked Out': days_out,
            'Expected Return': checkout.expected_return_date.strftime('%Y-%m-%d') if checkout.expected_return_date else 'Not set',
            'Assignment Type': checkout.assignment_type or 'checkout',
        })
    
    if format_type == 'csv':
        return generate_csv(data, 'long_term_checkouts.csv')
    elif format_type == 'excel':
        return generate_excel(data, 'long_term_checkouts.xlsx')
    elif format_type == 'pdf':
        return generate_pdf(data, 'long_term_checkouts.pdf', 'Long-Term Checkouts Report')
    else:
        return Response('Invalid format', status=400)


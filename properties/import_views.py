# properties/import_views.py
"""Import functionality for properties"""
from flask import render_template, request, redirect, url_for, flash, session
from flask_login import login_required, current_user
import csv
import io
from typing import List, Dict, Any
from werkzeug.utils import secure_filename

from utilities.tenant_helpers import tenant_query, tenant_add, tenant_commit, tenant_rollback
from middleware.tenant_middleware import tenant_required
from utilities.database import Property, PropertyUnit, utc_now, log_activity
from . import properties_bp


ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def parse_csv_file(file_content: str) -> tuple[List[str], List[Dict[str, str]]]:
    """Parse CSV file and return headers and rows"""
    reader = csv.DictReader(io.StringIO(file_content))
    headers = reader.fieldnames or []
    rows = [row for row in reader]
    return list(headers), rows


def parse_excel_file(file_bytes: bytes) -> tuple[List[str], List[Dict[str, str]]]:
    """Parse Excel file and return headers and rows"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel import")

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        return [], []

    headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(all_rows[0])]

    rows = []
    for row_data in all_rows[1:]:
        row_dict = {}
        for i, cell_value in enumerate(row_data):
            if i < len(headers):
                if cell_value is None:
                    row_dict[headers[i]] = ''
                else:
                    row_dict[headers[i]] = str(cell_value)
        rows.append(row_dict)

    wb.close()
    return headers, rows


PROPERTY_FIELDS = {
    'name': {'required': True, 'name': 'Property Name'},
    'type': {'required': False, 'name': 'Type', 'default': 'single_family'},
    'address_line1': {'required': True, 'name': 'Address Line 1'},
    'address_line2': {'required': False, 'name': 'Address Line 2'},
    'city': {'required': False, 'name': 'City'},
    'state': {'required': False, 'name': 'State'},
    'postal_code': {'required': False, 'name': 'Postal Code'},
    'country': {'required': False, 'name': 'Country', 'default': 'USA'},
    'notes': {'required': False, 'name': 'Notes'},
}

PROPERTY_UNIT_FIELDS = {
    'property_name': {'required': True, 'name': 'Property Name'},
    'label': {'required': True, 'name': 'Unit Label'},
    'floor': {'required': False, 'name': 'Floor'},
    'bedrooms': {'required': False, 'name': 'Bedrooms', 'type': 'int'},
    'bathrooms': {'required': False, 'name': 'Bathrooms', 'type': 'float'},
    'square_feet': {'required': False, 'name': 'Square Feet', 'type': 'int'},
    'notes': {'required': False, 'name': 'Notes'},
}


@properties_bp.route("/import", methods=["GET", "POST"])
@login_required
@tenant_required
def import_properties():
    """Import properties from CSV/Excel file"""
    if request.method == "POST":
        if 'file' not in request.files:
            flash("No file uploaded", "error")
            return redirect(request.url)

        file = request.files['file']

        if file.filename == '':
            flash("No file selected", "error")
            return redirect(request.url)

        if not allowed_file(file.filename):
            flash("Invalid file type. Please upload CSV or Excel file.", "error")
            return redirect(request.url)

        try:
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower()

            if file_ext == 'csv':
                file_content = file.read().decode('utf-8')
                headers, rows = parse_csv_file(file_content)
            else:
                file_bytes = file.read()
                headers, rows = parse_excel_file(file_bytes)

            if not rows:
                flash("No data found in file", "error")
                return redirect(request.url)

            # Store in session for next step
            session['import_data'] = {
                'headers': headers,
                'rows': rows,
                'filename': filename
            }

            return redirect(url_for('properties.import_properties_map'))

        except Exception as e:
            flash(f"Error reading file: {str(e)}", "error")
            return redirect(request.url)

    return render_template("properties/import_upload.html")


@properties_bp.route("/import/map", methods=["GET", "POST"])
@login_required
@tenant_required
def import_properties_map():
    """Map columns from uploaded file to database fields"""
    import_data = session.get('import_data')
    if not import_data:
        flash("No import data found. Please upload a file first.", "error")
        return redirect(url_for('properties.import_properties'))

    if request.method == "POST":
        mapping = {}
        for field in PROPERTY_FIELDS.keys():
            col = request.form.get(f'field_{field}', '').strip()
            if col:
                mapping[field] = col

        # Validate required fields are mapped
        missing = []
        for field, config in PROPERTY_FIELDS.items():
            if config.get('required') and field not in mapping:
                missing.append(config['name'])

        if missing:
            flash(f"Missing required fields: {', '.join(missing)}", "error")
            return render_template(
                "properties/import_map.html",
                headers=import_data['headers'],
                fields=PROPERTY_FIELDS,
                filename=import_data['filename']
            )

        session['import_mapping'] = mapping
        return redirect(url_for('properties.import_properties_process'))

    return render_template(
        "properties/import_map.html",
        headers=import_data['headers'],
        fields=PROPERTY_FIELDS,
        filename=import_data['filename']
    )


@properties_bp.route("/import/process", methods=["GET", "POST"])
@login_required
@tenant_required
def import_properties_process():
    """Process the import with mapped columns"""
    import_data = session.get('import_data')
    mapping = session.get('import_mapping')

    if not import_data or not mapping:
        flash("Import session expired. Please start over.", "error")
        return redirect(url_for('properties.import_properties'))

    if request.method == "POST":
        try:
            rows = import_data['rows']
            success_count = 0
            error_count = 0
            errors = []

            for idx, row in enumerate(rows, start=2):  # Start at 2 (after header row)
                try:
                    # Build property data from mapped columns
                    property_data = {}

                    for field, column in mapping.items():
                        value = row.get(column, '').strip()
                        if value:
                            property_data[field] = value
                        elif 'default' in PROPERTY_FIELDS[field]:
                            property_data[field] = PROPERTY_FIELDS[field]['default']

                    # Validate required fields
                    missing = []
                    for field, config in PROPERTY_FIELDS.items():
                        if config.get('required') and not property_data.get(field):
                            missing.append(config['name'])

                    if missing:
                        errors.append(f"Row {idx}: Missing {', '.join(missing)}")
                        error_count += 1
                        continue

                    # Check if property already exists (by name)
                    existing = tenant_query(Property).filter_by(name=property_data['name']).first()
                    if existing:
                        errors.append(f"Row {idx}: Property '{property_data['name']}' already exists")
                        error_count += 1
                        continue

                    # Create property
                    property_obj = Property(
                        name=property_data['name'],
                        type=property_data.get('type', 'single_family'),
                        address_line1=property_data['address_line1'],
                        address_line2=property_data.get('address_line2'),
                        city=property_data.get('city'),
                        state=property_data.get('state'),
                        postal_code=property_data.get('postal_code'),
                        country=property_data.get('country', 'USA'),
                        notes=property_data.get('notes')
                    )

                    tenant_add(property_obj)
                    success_count += 1

                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
                    error_count += 1
                    continue

            # Commit all changes
            tenant_commit()

            # Log activity
            log_activity(
                "properties_imported",
                user=current_user,
                summary=f"Imported {success_count} properties",
                meta={'success': success_count, 'errors': error_count},
                commit=True
            )

            # Clear session
            session.pop('import_data', None)
            session.pop('import_mapping', None)

            flash(f"Successfully imported {success_count} properties", "success")
            if error_count > 0:
                flash(f"{error_count} rows had errors", "warning")

            return render_template(
                "properties/import_result.html",
                success_count=success_count,
                error_count=error_count,
                errors=errors
            )

        except Exception as e:
            tenant_rollback()
            flash(f"Error during import: {str(e)}", "error")
            return redirect(url_for('properties.import_properties'))

    # GET - show preview
    rows = import_data['rows']
    preview_rows = rows[:10]  # Show first 10 rows

    # Build preview data
    preview_data = []
    for row in preview_rows:
        preview_item = {}
        for field, column in mapping.items():
            preview_item[PROPERTY_FIELDS[field]['name']] = row.get(column, '')
        preview_data.append(preview_item)

    return render_template(
        "properties/import_preview.html",
        preview_data=preview_data,
        total_rows=len(rows),
        fields=PROPERTY_FIELDS,
        mapping=mapping,
        filename=import_data['filename']
    )


# ==================== PROPERTY UNITS IMPORT ====================

@properties_bp.route("/units/import", methods=["GET", "POST"])
@login_required
@tenant_required
def import_units():
    """Import property units from CSV/Excel file"""
    if request.method == "POST":
        if 'file' not in request.files:
            flash("No file uploaded", "error")
            return redirect(request.url)

        file = request.files['file']

        if file.filename == '':
            flash("No file selected", "error")
            return redirect(request.url)

        if not allowed_file(file.filename):
            flash("Invalid file type. Please upload CSV or Excel file.", "error")
            return redirect(request.url)

        try:
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower()

            if file_ext == 'csv':
                file_content = file.read().decode('utf-8')
                headers, rows = parse_csv_file(file_content)
            else:
                file_bytes = file.read()
                headers, rows = parse_excel_file(file_bytes)

            if not rows:
                flash("No data found in file", "error")
                return redirect(request.url)

            session['import_data_units'] = {
                'headers': headers,
                'rows': rows,
                'filename': filename
            }

            return redirect(url_for('properties.import_units_map'))

        except Exception as e:
            flash(f"Error reading file: {str(e)}", "error")
            return redirect(request.url)

    return render_template("properties/import_units_upload.html")


@properties_bp.route("/units/import/map", methods=["GET", "POST"])
@login_required
@tenant_required
def import_units_map():
    """Map columns from uploaded file to database fields"""
    import_data = session.get('import_data_units')
    if not import_data:
        flash("No import data found. Please upload a file first.", "error")
        return redirect(url_for('properties.import_units'))

    if request.method == "POST":
        mapping = {}
        for field in PROPERTY_UNIT_FIELDS.keys():
            col = request.form.get(f'field_{field}', '').strip()
            if col:
                mapping[field] = col

        # Validate required fields are mapped
        missing = []
        for field, config in PROPERTY_UNIT_FIELDS.items():
            if config.get('required') and field not in mapping:
                missing.append(config['name'])

        if missing:
            flash(f"Missing required fields: {', '.join(missing)}", "error")
            return render_template(
                "properties/import_units_map.html",
                headers=import_data['headers'],
                fields=PROPERTY_UNIT_FIELDS,
                filename=import_data['filename']
            )

        session['import_mapping_units'] = mapping
        return redirect(url_for('properties.import_units_process'))

    return render_template(
        "properties/import_units_map.html",
        headers=import_data['headers'],
        fields=PROPERTY_UNIT_FIELDS,
        filename=import_data['filename']
    )


@properties_bp.route("/units/import/process", methods=["GET", "POST"])
@login_required
@tenant_required
def import_units_process():
    """Process the units import with mapped columns"""
    import_data = session.get('import_data_units')
    mapping = session.get('import_mapping_units')

    if not import_data or not mapping:
        flash("Import session expired. Please start over.", "error")
        return redirect(url_for('properties.import_units'))

    if request.method == "POST":
        try:
            rows = import_data['rows']
            success_count = 0
            error_count = 0
            errors = []

            for idx, row in enumerate(rows, start=2):
                try:
                    # Build unit data from mapped columns
                    unit_data = {}

                    for field, column in mapping.items():
                        value = row.get(column, '').strip()
                        if value:
                            # Handle type conversions
                            field_config = PROPERTY_UNIT_FIELDS[field]
                            if field_config.get('type') == 'int':
                                try:
                                    unit_data[field] = int(value)
                                except ValueError:
                                    errors.append(f"Row {idx}: Invalid integer for {field_config['name']}")
                                    error_count += 1
                                    continue
                            elif field_config.get('type') == 'float':
                                try:
                                    unit_data[field] = float(value)
                                except ValueError:
                                    errors.append(f"Row {idx}: Invalid number for {field_config['name']}")
                                    error_count += 1
                                    continue
                            else:
                                unit_data[field] = value

                    # Validate required fields
                    if not unit_data.get('property_name') or not unit_data.get('label'):
                        errors.append(f"Row {idx}: Missing Property Name or Unit Label")
                        error_count += 1
                        continue

                    # Find property by name
                    property_obj = tenant_query(Property).filter_by(name=unit_data['property_name']).first()
                    if not property_obj:
                        errors.append(f"Row {idx}: Property '{unit_data['property_name']}' not found")
                        error_count += 1
                        continue

                    # Check if unit already exists for this property
                    existing = tenant_query(PropertyUnit).filter_by(
                        property_id=property_obj.id,
                        label=unit_data['label']
                    ).first()

                    if existing:
                        errors.append(f"Row {idx}: Unit '{unit_data['label']}' already exists for property '{unit_data['property_name']}'")
                        error_count += 1
                        continue

                    # Create unit
                    unit = PropertyUnit(
                        property_id=property_obj.id,
                        label=unit_data['label'],
                        floor=unit_data.get('floor'),
                        bedrooms=unit_data.get('bedrooms'),
                        bathrooms=unit_data.get('bathrooms'),
                        square_feet=unit_data.get('square_feet'),
                        notes=unit_data.get('notes')
                    )

                    tenant_add(unit)
                    success_count += 1

                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
                    error_count += 1
                    continue

            # Commit all changes
            tenant_commit()

            # Log activity
            log_activity(
                "property_units_imported",
                user=current_user,
                summary=f"Imported {success_count} property units",
                meta={'success': success_count, 'errors': error_count},
                commit=True
            )

            # Clear session
            session.pop('import_data_units', None)
            session.pop('import_mapping_units', None)

            flash(f"Successfully imported {success_count} property units", "success")
            if error_count > 0:
                flash(f"{error_count} rows had errors", "warning")

            return render_template(
                "properties/import_units_result.html",
                success_count=success_count,
                error_count=error_count,
                errors=errors
            )

        except Exception as e:
            tenant_rollback()
            flash(f"Error during import: {str(e)}", "error")
            return redirect(url_for('properties.import_units'))

    # GET - show preview
    rows = import_data['rows']
    preview_rows = rows[:10]

    # Build preview data
    preview_data = []
    for row in preview_rows:
        preview_item = {}
        for field, column in mapping.items():
            preview_item[PROPERTY_UNIT_FIELDS[field]['name']] = row.get(column, '')
        preview_data.append(preview_item)

    return render_template(
        "properties/import_units_preview.html",
        preview_data=preview_data,
        total_rows=len(rows),
        fields=PROPERTY_UNIT_FIELDS,
        mapping=mapping,
        filename=import_data['filename']
    )
